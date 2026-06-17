"""
Вкладка «Генерация» — основной рабочий экран.
"""
import threading
import logging
import queue
from datetime import datetime
from pathlib import Path

import yaml
import customtkinter as ctk
from tkinter import filedialog

logger = logging.getLogger(__name__)



class GenerationView(ctk.CTkFrame):
    def __init__(self, parent, navigate_to=None, get_presets_view=None):
        super().__init__(parent, corner_radius=0, fg_color="transparent")
        self._navigate_to      = navigate_to
        self._get_presets_view = get_presets_view

        self._preset_data: dict | None = None
        self._input_file:  Path | None = None
        self._km_file:     Path | None = None
        self._output_fmt   = ctk.StringVar(value="1")
        self._mode_var     = ctk.StringVar(value="1")
        self._use_km_var   = ctk.BooleanVar(value=False)
        self._no_container = ctk.BooleanVar(value=True)
        self._log_queue: queue.Queue = queue.Queue()

        # Шрифты создаются здесь — после создания root window
        self._lbl_font   = ctk.CTkFont(size=11, weight="bold")
        self._body_font  = ctk.CTkFont(size=12)
        self._small_font = ctk.CTkFont(size=11)

        self._build()
        self._load_presets()

    # ------------------------------------------------------------------
    def _section(self, parent, title: str) -> ctk.CTkFrame:
        outer = ctk.CTkFrame(parent, corner_radius=10)
        outer.pack(fill="x", padx=16, pady=6)
        ctk.CTkLabel(outer, text=title, font=self._lbl_font, text_color="gray60").pack(
            anchor="w", padx=12, pady=(8, 2)
        )
        return outer

    # ------------------------------------------------------------------
    # Построение UI
    # ------------------------------------------------------------------
    def _build(self):
        # Скроллируемая область для всех секций
        scroll = ctk.CTkScrollableFrame(self, corner_radius=0, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        self._build_preset_section(scroll)
        self._build_file_section(scroll)
        self._build_params_section(scroll)
        self._build_mode_section(scroll)
        self._build_km_section(scroll)
        self._build_format_section(scroll)
        self._build_inject_section(scroll)

        # Кнопка генерации
        self._generate_btn = ctk.CTkButton(
            scroll,
            text="⚡  СГЕНЕРИРОВАТЬ ДОКУМЕНТЫ",
            height=46,
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=10,
            command=self._on_generate,
        )
        self._generate_btn.pack(fill="x", padx=16, pady=(10, 4))

        # Прогресс + лог результата
        self._build_progress_section(scroll)

    # ── 1. Клиент ──────────────────────────────────────────────────────
    def _build_preset_section(self, parent):
        sec = self._section(parent, "1.  КЛИЕНТ")

        row = ctk.CTkFrame(sec, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(0, 4))

        self._preset_var      = ctk.StringVar()
        self._preset_dropdown = ctk.CTkOptionMenu(
            row,
            variable=self._preset_var,
            values=["—"],
            command=self._on_preset_selected,
            width=220,
            font=self._body_font,
        )
        self._preset_dropdown.pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            row, text="✏️  Редактировать", width=130, font=self._body_font,
            fg_color="transparent", border_width=1,
            command=self._on_edit_preset,
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row, text="+ Создать", width=100, font=self._body_font,
            command=self._on_create_preset,
        ).pack(side="left")

        self._preset_info_lbl = ctk.CTkLabel(
            sec, text="", font=self._small_font, text_color="gray55", justify="left"
        )
        self._preset_info_lbl.pack(anchor="w", padx=14, pady=(0, 8))

    # ── 2. Входной файл ────────────────────────────────────────────────
    def _build_file_section(self, parent):
        sec = self._section(parent, "2.  ВХОДНОЙ ФАЙЛ")
        row = ctk.CTkFrame(sec, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(0, 4))

        ctk.CTkButton(
            row, text="📂  Из папки input/", width=150, font=self._body_font,
            command=self._pick_from_input,
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            row, text="Обзор...", width=90, fg_color="transparent",
            border_width=1, font=self._body_font, command=self._pick_file_dialog,
        ).pack(side="left")

        self._file_lbl = ctk.CTkLabel(
            sec, text="Файл не выбран", font=self._small_font, text_color="gray55"
        )
        self._file_lbl.pack(anchor="w", padx=14, pady=(2, 8))

    # ── 3. Параметры инвойса ────────────────────────────────────────────
    def _build_params_section(self, parent):
        sec = self._section(parent, "3.  ПАРАМЕТРЫ ИНВОЙСА")
        grid = ctk.CTkFrame(sec, fg_color="transparent")
        grid.pack(fill="x", padx=10, pady=(0, 10))
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(3, weight=1)

        ctk.CTkLabel(grid, text="Номер инвойса:", font=self._body_font).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        self._invoice_entry = ctk.CTkEntry(grid, font=self._body_font, placeholder_text=datetime.now().strftime("%Y%m%d01"))
        self._invoice_entry.grid(row=0, column=1, sticky="ew", padx=(0, 20), pady=4)
        self._invoice_entry.insert(0, datetime.now().strftime("%Y%m%d01"))

        ctk.CTkLabel(grid, text="Контейнер:", font=self._body_font).grid(row=0, column=2, sticky="w", padx=(0, 8), pady=4)
        self._container_entry = ctk.CTkEntry(grid, font=self._body_font, placeholder_text="TCKU1234567")
        self._container_entry.grid(row=0, column=3, sticky="ew", pady=4)

        ctk.CTkCheckBox(
            grid, text="Нет контейнера", variable=self._no_container,
            font=self._small_font, command=self._toggle_container,
        ).grid(row=1, column=2, columnspan=2, sticky="w", pady=(0, 4))
        self._toggle_container()

    # ── 4. Режим ────────────────────────────────────────────────────────
    def _build_mode_section(self, parent):
        sec = self._section(parent, "4.  РЕЖИМ ОБРАБОТКИ")
        row = ctk.CTkFrame(sec, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkRadioButton(
            row, text="🚢  Загрузка контейнера (пары)", variable=self._mode_var,
            value="1", font=self._body_font,
        ).pack(side="left", padx=(0, 24))

        ctk.CTkRadioButton(
            row, text="📦  Отгрузка (полупары)", variable=self._mode_var,
            value="2", font=self._body_font,
        ).pack(side="left")

    # ── 5. Маркировка КМ ────────────────────────────────────────────────
    def _build_km_section(self, parent):
        sec = self._section(parent, "5.  МАРКИРОВКА КМ  (Честный знак)")
        top = ctk.CTkFrame(sec, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=(0, 4))

        ctk.CTkCheckBox(
            top, text="Добавить КМ коды в Спецификацию",
            variable=self._use_km_var, font=self._body_font,
            command=self._toggle_km,
        ).pack(side="left")

        self._km_frame = ctk.CTkFrame(sec, fg_color="transparent")
        self._km_frame.pack(fill="x", padx=10, pady=(0, 8))

        self._km_dropdown = ctk.CTkOptionMenu(
            self._km_frame, font=self._body_font, width=280,
            command=self._on_km_selected,
        )
        self._km_dropdown.pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            self._km_frame, text="↻", width=32, font=self._body_font,
            command=self._refresh_km_list,
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            self._km_frame, text="📂  Обзор...", width=110, font=self._body_font,
            fg_color="transparent", border_width=1,
            command=self._browse_km_file,
        ).pack(side="left")

        self._toggle_km()

    # ── 6. Формат вывода ────────────────────────────────────────────────
    def _build_format_section(self, parent):
        sec = self._section(parent, "6.  ФОРМАТ ВЫВОДА")
        row = ctk.CTkFrame(sec, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkRadioButton(
            row, text="Три отдельных файла  (.xlsx × 3)",
            variable=self._output_fmt, value="1", font=self._body_font,
        ).pack(side="left", padx=(0, 24))

        ctk.CTkRadioButton(
            row, text="Один файл с тремя листами  (.xlsx)",
            variable=self._output_fmt, value="2", font=self._body_font,
        ).pack(side="left")

    # ── 7. Добавить КИЗ в готовый файл ─────────────────────────────────
    def _build_inject_section(self, parent):
        sec = self._section(parent, "7.  ДОБАВИТЬ КИЗ В ГОТОВЫЙ ФАЙЛ")
        row = ctk.CTkFrame(sec, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(
            row,
            text="Если документы уже созданы, но тогда ещё не было кодов КМ:",
            font=self._small_font, text_color="gray55",
        ).pack(anchor="w", pady=(0, 6))

        ctk.CTkButton(
            row, text="🏷️  Добавить КИЗ в существующий файл...",
            height=36, font=self._body_font, fg_color=("gray75", "gray30"),
            hover_color=("gray65", "gray40"),
            command=self._on_inject_kiz,
        ).pack(anchor="w")

    # ── Прогресс / результат ────────────────────────────────────────────
    def _build_progress_section(self, parent):
        self._progress_frame = ctk.CTkFrame(parent, corner_radius=10)
        self._progress_frame.pack(fill="x", padx=16, pady=(4, 16))
        self._progress_frame.pack_forget()  # скрыт по умолчанию

        self._progress_bar = ctk.CTkProgressBar(self._progress_frame, mode="indeterminate")
        self._progress_bar.pack(fill="x", padx=12, pady=(10, 4))

        self._progress_log = ctk.CTkTextbox(
            self._progress_frame, height=140, font=ctk.CTkFont(family="Consolas", size=11),
            state="disabled",
        )
        self._progress_log.pack(fill="x", padx=12, pady=(0, 10))

        self._open_btn = ctk.CTkButton(
            self._progress_frame, text="📁  Открыть папку output/",
            height=34, font=self._body_font,
            command=lambda: self._open_output_folder(),
        )
        self._open_btn.pack(pady=(0, 10))
        self._open_btn.pack_forget()

    # ------------------------------------------------------------------
    # Загрузка пресетов
    # ------------------------------------------------------------------
    def _load_presets(self, keep_selection: str = None):
        """Загружает пресеты и обновляет dropdown."""
        presets_dir = Path("presets")
        self._presets: list[dict] = []

        if presets_dir.exists():
            for f in sorted(presets_dir.glob("*.yaml")):
                try:
                    data = yaml.safe_load(f.read_text(encoding="utf-8"))
                    self._presets.append({"file": f, "name": data.get("preset_name", f.stem), "data": data})
                except Exception:
                    pass

        if not self._presets:
            self._preset_dropdown.configure(values=["— нет пресетов —"])
            self._preset_var.set("— нет пресетов —")
            self._preset_data = None
            self._preset_info_lbl.configure(text="Создайте пресет кнопкой справа")
            return

        names = [p["name"] for p in self._presets]
        self._preset_dropdown.configure(values=names)

        # Восстанавливаем выбор или берём первый
        if keep_selection and keep_selection in names:
            self._preset_var.set(keep_selection)
        else:
            self._preset_var.set(names[0])

        self._on_preset_selected(self._preset_var.get())

    def _on_preset_selected(self, name: str):
        match = next((p for p in self._presets if p["name"] == name), None)
        if not match:
            return
        self._preset_data = match["data"]
        p = self._preset_data
        seller   = p.get("seller",   {}).get("name",  "—")
        buyer    = p.get("buyer",    {}).get("name",  "—")
        contract = p.get("contract", {}).get("number","—")
        self._preset_info_lbl.configure(
            text=f"Продавец: {seller}   |   Покупатель: {buyer}   |   Договор: {contract}"
        )

    def _on_edit_preset(self):
        pv = self._get_presets_view() if self._get_presets_view else None
        if pv:
            current_name = self._preset_var.get()
            idx = next((i for i, p in enumerate(pv._presets) if p["data"].get("preset_name") == current_name), 0)
            pv._load_presets()
            pv._select(idx)
        if self._navigate_to:
            self._navigate_to("presets")

    def _on_create_preset(self):
        pv = self._get_presets_view() if self._get_presets_view else None
        if pv:
            pv._new_preset()
        if self._navigate_to:
            self._navigate_to("presets")

    # ------------------------------------------------------------------
    # Выбор входного файла
    # ------------------------------------------------------------------
    def _pick_from_input(self):
        input_dir = Path("input")
        if not input_dir.exists():
            self._file_lbl.configure(text="Папка input/ не найдена", text_color="red")
            return

        files = sorted(
            [f for f in input_dir.glob("*.xlsx") if not f.name.startswith("~")],
            key=lambda f: f.stat().st_mtime, reverse=True,
        )
        if not files:
            self._file_lbl.configure(text="Нет Excel файлов в input/", text_color="orange")
            return

        # Если файл один — берём сразу
        if len(files) == 1:
            self._set_input_file(files[0])
            return

        # Иначе — диалог с кратким списком
        win = ctk.CTkToplevel(self)
        win.title("Выберите файл из input/")
        win.geometry("500x320")
        win.grab_set()

        ctk.CTkLabel(win, text="Файлы в папке input/:", font=self._lbl_font).pack(pady=(12, 4), padx=16, anchor="w")

        listbox = ctk.CTkScrollableFrame(win)
        listbox.pack(fill="both", expand=True, padx=16, pady=4)

        for f in files:
            size_kb = f.stat().st_size / 1024
            mtime   = datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
            row = ctk.CTkFrame(listbox, corner_radius=6, cursor="hand2")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f.name, font=self._body_font, anchor="w").pack(side="left", padx=10, pady=6)
            ctk.CTkLabel(row, text=f"{size_kb:.0f} KB  {mtime}", font=self._small_font, text_color="gray55").pack(side="right", padx=10)
            row.bind("<Button-1>", lambda e, file=f, w=win: (self._set_input_file(file), w.destroy()))
            for child in row.winfo_children():
                child.bind("<Button-1>", lambda e, file=f, w=win: (self._set_input_file(file), w.destroy()))

    def _pick_file_dialog(self):
        path = filedialog.askopenfilename(
            title="Выберите входной Excel файл",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=str(Path("input")) if Path("input").exists() else ".",
        )
        if path:
            self._set_input_file(Path(path))

    def _set_input_file(self, f: Path):
        self._input_file = f
        size_kb = f.stat().st_size / 1024
        mtime   = datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y")
        self._file_lbl.configure(
            text=f"✅  {f.name}  ({size_kb:.0f} KB, {mtime})",
            text_color=("gray20", "gray85"),
        )

    # ------------------------------------------------------------------
    # Контейнер / КМ
    # ------------------------------------------------------------------
    def _toggle_container(self):
        if self._no_container.get():
            self._container_entry.configure(state="disabled", placeholder_text="нет")
            self._container_entry.delete(0, "end")
        else:
            self._container_entry.configure(state="normal", placeholder_text="TCKU1234567")

    def _toggle_km(self):
        if self._use_km_var.get():
            self._km_frame.pack(fill="x", padx=10, pady=(0, 8))
            self._refresh_km_list()
        else:
            self._km_frame.pack_forget()
            self._km_file = None

    def _refresh_km_list(self):
        km_dir = Path("выгрузка честный знак")
        files = []
        if km_dir.exists():
            files = sorted(
                [f for f in km_dir.glob("*.xlsx") if not f.name.startswith("~")],
                key=lambda f: f.stat().st_mtime, reverse=True,
            )
        if not hasattr(self, "_km_files_map"):
            self._km_files_map = {}
        if files:
            names = [f.name for f in files]
            self._km_files_map.update({f.name: f for f in files})
            # Добавляем новые имена в dropdown, не затирая вручную добавленные
            current_values = list(self._km_dropdown.cget("values") or [])
            merged = list(dict.fromkeys(names + [v for v in current_values if v not in names and v != "— нет файлов —"]))
            self._km_dropdown.configure(values=merged)
            self._km_dropdown.set(merged[0])
            self._km_file = files[0]
        else:
            if not self._km_files_map:
                self._km_dropdown.configure(values=["— нет файлов —"])
                self._km_dropdown.set("— нет файлов —")
                self._km_file = None

    def _browse_km_file(self):
        path = filedialog.askopenfilename(
            title="Выберите файл маркировки (Честный знак)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        f = Path(path)
        if not hasattr(self, "_km_files_map"):
            self._km_files_map = {}
        self._km_files_map[f.name] = f
        current_values = list(self._km_dropdown.cget("values") or [])
        if f.name not in current_values:
            current_values = [v for v in current_values if v != "— нет файлов —"]
            current_values.insert(0, f.name)
            self._km_dropdown.configure(values=current_values)
        self._km_dropdown.set(f.name)
        self._km_file = f

    def _on_km_selected(self, value: str):
        self._km_file = self._km_files_map.get(value)

    # ------------------------------------------------------------------
    # Генерация документов
    # ------------------------------------------------------------------
    def _on_generate(self):
        if not self._validate():
            return

        invoice_number   = self._invoice_entry.get().strip()
        container_number = "" if self._no_container.get() else self._container_entry.get().strip()
        mode             = "container" if self._mode_var.get() == "1" else "shipment"
        output_fmt       = self._output_fmt.get()
        km_file          = self._km_file if self._use_km_var.get() else None

        self._set_processing(True)

        thread = threading.Thread(
            target=self._run_generation,
            args=(self._input_file, self._preset_data, invoice_number, container_number, mode, output_fmt, km_file),
            daemon=True,
        )
        thread.start()
        self._poll_log_queue()

    def _validate(self) -> bool:
        if not self._preset_data:
            self._show_error("Выберите пресет клиента.")
            return False
        if not self._input_file or not self._input_file.exists():
            self._show_error("Выберите входной файл.")
            return False
        if not self._invoice_entry.get().strip():
            self._show_error("Введите номер инвойса.")
            return False
        return True

    def _show_error(self, msg: str):
        win = ctk.CTkToplevel(self)
        win.title("Ошибка")
        win.geometry("380x140")
        win.grab_set()
        ctk.CTkLabel(win, text=msg, font=self._body_font, wraplength=340).pack(expand=True, pady=20)
        ctk.CTkButton(win, text="OK", command=win.destroy, width=80).pack(pady=(0, 16))

    def _set_processing(self, active: bool):
        self._generate_btn.configure(state="disabled" if active else "normal")
        if active:
            self._progress_frame.pack(fill="x", padx=16, pady=(4, 16))
            self._open_btn.pack_forget()
            self._clear_log()
            self._progress_bar.start()
        else:
            self._progress_bar.stop()
            self._progress_bar.set(1.0)

    def _run_generation(self, input_file, preset, invoice_number, container_number, mode, output_fmt, km_file):
        """Запускается в фоновом потоке."""
        import yaml as _yaml
        from src.parser            import InputFileParser
        from src.processor         import DataProcessor
        from src.shipment_parser   import ShipmentParser
        from src.shipment_processor import ShipmentProcessor
        from src.generators.invoice       import InvoiceGenerator
        from src.generators.specification import SpecificationGenerator
        from src.generators.packing_list  import PackingListGenerator
        from src.models    import DocumentMetadata
        from src.km_loader import KMLoader
        import tempfile, os
        from openpyxl import Workbook, load_workbook
        from copy import copy

        def log(msg): self._log_queue.put(("msg", msg))
        def err(msg): self._log_queue.put(("err", msg))

        try:
            config_path = Path("config.yaml")
            with open(config_path, encoding="utf-8") as f:
                config = _yaml.safe_load(f)

            # 1. Парсинг
            log(f"📥  Чтение файла: {input_file.name}")
            if mode == "container":
                parser   = InputFileParser(config)
                products = parser.parse(str(input_file))
                if not products:
                    err("Не найдено ни одной позиции."); return
                log(f"✅  Загружено {len(products)} позиций")
                processor    = DataProcessor(config)
                output_lines = processor.process(products)
            else:
                parser        = ShipmentParser(str(input_file))
                shipment_lines = parser.parse()
                if not shipment_lines:
                    err("Не найдено ни одной позиции."); return
                log(f"✅  Загружено {len(shipment_lines)} строк (полупар)")
                processor    = ShipmentProcessor(config)
                output_lines = processor.process(shipment_lines)

            log(f"⚙️   Создано {len(output_lines)} строк для документов")

            # 2. КМ коды
            if km_file:
                log(f"🏷️   Загрузка КМ кодов из {km_file.name}...")
                km_loader = KMLoader(str(km_file))
                count = enriched = 0
                for line in output_lines:
                    if line.qty_by_size:
                        codes = km_loader.get_km_codes_exact(line.article, line.qty_by_size)
                        if codes:
                            line.kiz_codes = codes
                            enriched += 1
                            count    += len(codes)
                log(f"✅  КМ: {count} кодов в {enriched} строках")

            # 3. Метаданные
            date     = datetime.now().strftime("%d.%m.%Y")
            metadata = DocumentMetadata(
                invoice_number   = invoice_number,
                date             = date,
                container_number = container_number,
                seller_name      = preset["seller"]["name"],
                seller_name_en   = preset["seller"].get("name_en", ""),
                seller_address   = preset["seller"]["address"],
                seller_address_en= preset["seller"].get("address_en", ""),
                buyer_name       = preset["buyer"]["name"],
                buyer_address    = preset["buyer"]["address"],
                buyer_address_en = preset["buyer"].get("address_en", ""),
                contract_number  = preset["contract"]["number"],
                contract_date    = preset["contract"]["date"],
                terms_of_delivery= preset["delivery"]["terms"],
                currency         = preset["delivery"].get("currency", "CNY"),
            )

            # 4. Генерация
            output_path = Path("output")
            output_path.mkdir(exist_ok=True)
            base = f"Shusteri_{invoice_number}_{datetime.now().strftime('%Y%m%d')}"

            log("📄  Генерация Invoice...")
            if output_fmt == "1":
                inv_path     = output_path / f"{base}_Invoice.xlsx"
                spec_path    = output_path / f"{base}_Specification.xlsx"
                packing_path = output_path / f"{base}_PackingList.xlsx"

                InvoiceGenerator(config, preset, mode=mode).generate(output_lines, metadata, str(inv_path))
                log("📄  Генерация Specification...")
                SpecificationGenerator(config, preset, mode=mode).generate(output_lines, metadata, str(spec_path))
                log("📄  Генерация Packing List...")
                PackingListGenerator(config, preset, mode=mode).generate(output_lines, metadata, str(packing_path))

                files = [inv_path, spec_path, packing_path]
            else:
                log("📄  Сборка единого файла (3 листа)...")
                combined = output_path / f"{base}_All_Documents.xlsx"
                with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
                    ti = os.path.join(tmp, "inv.xlsx")
                    ts = os.path.join(tmp, "spec.xlsx")
                    tp = os.path.join(tmp, "pack.xlsx")
                    InvoiceGenerator(config, preset, mode=mode).generate(output_lines, metadata, ti)
                    SpecificationGenerator(config, preset, mode=mode).generate(output_lines, metadata, ts)
                    PackingListGenerator(config, preset, mode=mode).generate(output_lines, metadata, tp)

                    wb_out = Workbook()
                    wb_out.remove(wb_out.active)
                    for sheet_name, tmp_path in [("Invoice", ti), ("Specification", ts), ("Packing List", tp)]:
                        wb_src = load_workbook(tmp_path)
                        ws_src = wb_src.active
                        ws_dst = wb_out.create_sheet(sheet_name)
                        for col in ws_src.column_dimensions:
                            ws_dst.column_dimensions[col].width = ws_src.column_dimensions[col].width
                        for row_dim in ws_src.row_dimensions:
                            ws_dst.row_dimensions[row_dim].height = ws_src.row_dimensions[row_dim].height
                        for row in ws_src.iter_rows():
                            for cell in row:
                                tc = ws_dst[cell.coordinate]
                                tc.value = cell.value
                                if cell.has_style:
                                    tc.font = copy(cell.font); tc.border = copy(cell.border)
                                    tc.fill = copy(cell.fill); tc.alignment = copy(cell.alignment)
                                    tc.number_format = copy(cell.number_format)
                        for mr in ws_src.merged_cells.ranges:
                            ws_dst.merge_cells(str(mr))
                        wb_src.close()
                    wb_out.save(str(combined))
                    wb_out.close()
                files = [combined]

            # Статистика
            total_qty    = sum(l.quantity    for l in output_lines)
            total_amount = sum(l.amount      for l in output_lines)
            total_net    = sum(l.net_weight  for l in output_lines)
            total_gross  = sum(l.gross_weight for l in output_lines)

            log("")
            log(f"✅  Готово! Создано файлов: {len(files)}")
            for f in files:
                log(f"   📎  {f.name}")
            log(f"")
            log(f"   Пар: {total_qty}   |   Сумма: ¥{total_amount:,.2f}")
            log(f"   Нетто: {total_net:.3f} кг   |   Брутто: {total_gross:.3f} кг")

            self._log_queue.put(("done", str(output_path.absolute())))

        except Exception as e:
            logger.exception("Ошибка генерации")
            err(f"❌  Ошибка: {e}")
            self._log_queue.put(("fail", None))

    # ------------------------------------------------------------------
    # Инъекция КИЗ в готовый файл
    # ------------------------------------------------------------------
    def _on_inject_kiz(self):
        from ui.views._inject_dialog import InjectKIZDialog
        InjectKIZDialog(self)

    # ------------------------------------------------------------------
    # Очередь лога (polling из UI-потока)
    # ------------------------------------------------------------------
    def _poll_log_queue(self):
        try:
            while True:
                kind, data = self._log_queue.get_nowait()
                if kind == "msg":
                    self._append_log(data, color=None)
                elif kind == "err":
                    self._append_log(data, color="red")
                elif kind == "done":
                    self._output_folder = data
                    self._set_processing(False)
                    self._open_btn.pack(pady=(0, 10))
                    self._show_success_toast(data)
                    return
                elif kind == "fail":
                    self._set_processing(False)
                    return
        except queue.Empty:
            pass
        self.after(100, self._poll_log_queue)

    def _show_success_toast(self, output_folder: str):
        """Небольшое всплывающее уведомление об успешной генерации."""
        toast = ctk.CTkToplevel(self)
        toast.title("Готово!")
        toast.geometry("420x200")
        toast.resizable(False, False)
        toast.grab_set()

        # Центрируем относительно главного окна
        self.update_idletasks()
        rx = self.winfo_rootx() + self.winfo_width()  // 2 - 210
        ry = self.winfo_rooty() + self.winfo_height() // 2 - 100
        toast.geometry(f"+{rx}+{ry}")

        ctk.CTkLabel(
            toast, text="✅  Документы успешно созданы!",
            font=ctk.CTkFont(size=14, weight="bold"),
        ).pack(pady=(22, 6))

        ctk.CTkLabel(
            toast, text=f"Папка: {output_folder}",
            font=ctk.CTkFont(size=11), text_color="gray55", wraplength=380,
        ).pack(pady=(0, 16))

        btn_row = ctk.CTkFrame(toast, fg_color="transparent")
        btn_row.pack()

        import subprocess, sys
        ctk.CTkButton(
            btn_row, text="📁  Открыть папку", width=140,
            font=ctk.CTkFont(size=12),
            command=lambda: (
                subprocess.Popen(["explorer", output_folder]) if sys.platform == "win32" else None,
                toast.destroy(),
            ),
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_row, text="Закрыть", width=90,
            font=ctk.CTkFont(size=12),
            fg_color="transparent", border_width=1,
            command=toast.destroy,
        ).pack(side="left")

        # Авто-закрытие через 8 секунд
        toast.after(8000, lambda: toast.destroy() if toast.winfo_exists() else None)

    def _append_log(self, text: str, color=None):
        self._progress_log.configure(state="normal")
        self._progress_log.insert("end", text + "\n")
        self._progress_log.see("end")
        self._progress_log.configure(state="disabled")

    def _clear_log(self):
        self._progress_log.configure(state="normal")
        self._progress_log.delete("1.0", "end")
        self._progress_log.configure(state="disabled")

    def _open_output_folder(self):
        import subprocess, sys
        folder = getattr(self, "_output_folder", str(Path("output").absolute()))
        if sys.platform == "win32":
            subprocess.Popen(["explorer", folder])
