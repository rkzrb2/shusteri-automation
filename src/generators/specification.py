"""
Генератор Specification БЕЗ группировки
Выводит все строки из output_lines подряд
"""
import pandas as pd
from typing import List
from ..models import OutputLine, DocumentMetadata


class SpecificationGenerator:
    """Генератор Спецификации"""

    def __init__(self, config: dict, preset: dict, mode: str = 'container'):
        """
        Args:
            config: конфигурация
            preset: пресет клиента
            mode: режим работы ('container' или 'shipment')
        """
        self.config = config
        self.preset = preset
        self.mode = mode

    def generate(
            self,
            lines: List[OutputLine],
            metadata: DocumentMetadata,
            output_path: str
    ) -> str:
        """Генерирует Specification"""

        rows = []

        # Шапка (правка 8 - текст "Спецификация к контракту" справа)
        rows.append(['', '', '', '', '', '', '', '', '', '', '', '', 'Спецификация к Контракту/', '', '', '', '', ''])
        rows.append(['', '', '', '', '', '', '', '', '', '', '', '', 'Specification to the Contract', '', '', '', '', ''])
        rows.append(['', '', '', '', '', '', '', '', '', '', '', '', f'№{metadata.contract_number} from/от {metadata.contract_date}', '', '', '', '', ''])

        rows.append(['', '', '', '', '', '', '', '', '', '', '', '', f'Container / Контейнер № {metadata.container_number}', '', '', '', '', ''])

        rows.append([''])  # Пустая строка 5
        rows.append([''])  # Пустая строка 6
        rows.append([''])  # Пустая строка 7
        rows.append([f'Specification / Спецификация № {metadata.invoice_number} from/от {metadata.date}'])

        rows.append([''])  # Пустая строка 9

        header = [
            "№",
            "Brand / Марка",
            "Code / Код ТНВЭД",
            "Factory code / Артикул",
            " Description / Описание",
            "Color / Цвет",
            "Top material / Материал верха",
            "Lining material / Материал подкладки",
            "Outsole material / Материал подошвы",
            "Heel height / Высота каблука",
            "Insole length / Длина стельки",
            "Quantity of pairs / Количество пар",
            "Net weight (kg) / Вес нетто (кг)",
            "Gross weight (kg) / Вес брутто (кг)",
            "Quantity of places / Количество мест",
            "Price / Цена, cny",
            "Amount / Сумма, cny",
            "КИЗ"
        ]
        rows.append(header)

        # КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: БЕЗ ГРУППИРОВКИ - просто все строки подряд!
        total_qty = 0
        total_net = 0
        total_gross = 0
        total_boxes = 0
        total_amount = 0
        
        data_start_row = 11  # Строка, где начинаются данные
        
        # Отслеживаем номер позиции и пары строк
        item_number = 0
        prev_line = None

        for idx, line in enumerate(lines):
            # Определяем является ли строка "второй частью" артикула
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )
            
            # Увеличиваем номер только для первых строк
            if not is_continuation:
                item_number += 1

            description = '' if is_continuation else line.description

            rows.append([
                '' if is_continuation else item_number,
                '' if is_continuation else line.brand,
                line.hs_code,
                '' if is_continuation else line.article,
                description,
                '' if is_continuation else line.color,
                '' if is_continuation else line.material,
                '' if is_continuation else line.lining,
                '' if is_continuation else line.sole,
                '' if is_continuation else line.heel_height,
                line.insole_category,
                line.quantity,
                float(line.net_weight),
                float(line.gross_weight),
                '',  # Boxes - будет заполнено после объединения
                float(line.price),
                float(line.amount),
                ' '.join(line.kiz_codes) if line.kiz_codes else ''
            ])

            total_qty += line.quantity
            total_net += line.net_weight
            total_gross += line.gross_weight
            # ИСПРАВЛЕНО: Правильный подсчет boxes
            # Для второй строки пары (is_continuation) НЕ добавляем boxes,
            # так как они уже были добавлены в первой строке
            if not is_continuation:
                total_boxes += line.boxes
            total_amount += line.amount
            
            prev_line = line

        data_end_row = len(rows)  # Последняя строка с данными

        # Итоги
        rows.append(['', '', '', '', 'Total / Итого:', '', '', '', '', '', '', total_qty, round(float(total_net), 3), round(float(total_gross), 3), total_boxes, '', float(total_amount), ''])

        # Футер с подписями
        rows.append([f"-Manufacturer / Производитель: {metadata.seller_name}"])
        rows.append([f"-Country of origin / Страна происхождения: {self.preset['delivery']['country_of_origin_en']} / {self.preset['delivery']['country_of_origin']}"])
        rows.append(["-Country of destination / Страна назначения: Russia / Россия"])
        rows.append(["-Product not for military use / Товар не для применения в военных целях"])
        rows.append([f"-Terms of delivery / Условия поставки: {metadata.terms_of_delivery}"])
        rows.append(["-Terms of payment / Условия оплаты:"])
        rows.append([f"Payment of the cost of this transaction for the delivery of the goods specified above in the framework of the execution of Contract No. {metadata.contract_number} dated {metadata.contract_date} in the amount of ¥ {total_amount:,.2f} is payable no later than 120 days from the date of filing the Declaration for the goods in the country of Import./ Оплата стоимости данной сделки по поставке товара, указанного выше в рамках исполнения Контракта № {metadata.contract_number} от {metadata.contract_date} г. в размере ¥ {total_amount:,.2f} подлежит оплате не позднее 120 дней с даты подачи Декларации на товар в стране Импорта"])

        # Подписи (только в Specification!)
        rows.append(['', 'Buyer / Покупатель:', '', '', '', '', '', '', '', '', 'Seller / Продавец:', '', '', '', '', '', '', ''])
        rows.append(['', metadata.buyer_name.split('/')[0].strip(), '', '', '', '', '', '', '', '', metadata.seller_name, '', '', '', '', '', '', ''])

        # Создаем DataFrame и сохраняем
        df = pd.DataFrame(rows)
        # ВАЖНО: Используем контекстный менеджер чтобы файл корректно закрылся
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        
        # Объединяем ячейки boxes для пар строк
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Объединяем ячейки boxes для пар строк
        self._merge_cells_container(ws, lines, data_start_row)

        wb.save(output_path)
        wb.close()

        # Применяем форматирование
        from ..formatters import SpecificationFormatter
        formatter = SpecificationFormatter(self.config)
        formatter.format(output_path, data_start_row, data_end_row)

        return output_path

    def _merge_cells_container(self, ws, lines: List[OutputLine], data_start_row: int):
        """
        Объединяет ячейки boxes для пар строк с одним артикулом (≤24см и >24см)

        Колонка O = Qty of places (коробки)
        """
        prev_line = None
        current_row = data_start_row

        for line in lines:
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )

            if is_continuation:
                # Объединяем ячейки boxes (колонка O) для предыдущей и текущей строки
                ws.merge_cells(f'O{current_row-1}:O{current_row}')
                # Записываем ИСХОДНОЕ значение boxes
                ws[f'O{current_row-1}'] = line.original_boxes
            elif line.boxes > 0:
                # Для обычных строк просто записываем значение
                ws[f'O{current_row}'] = line.boxes

            current_row += 1
            prev_line = line
