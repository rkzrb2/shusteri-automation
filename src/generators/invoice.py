"""
Генератор Invoice БЕЗ группировки
Выводит все строки из output_lines подряд
"""
import pandas as pd
from typing import List
from ..models import OutputLine, DocumentMetadata


class InvoiceGenerator:
    """Генератор Инвойса"""

    def __init__(self, config: dict, preset: dict, mode: str = 'container'):
        """
        Args:
            config: конфигурация
            preset: пресет клиента
            mode: режим работы ('container' или 'shipment')
        """
        self.config = config
        self.preset = preset
        self.mode = mode

    def generate(
            self,
            lines: List[OutputLine],
            metadata: DocumentMetadata,
            output_path: str
    ) -> str:
        """
        Генерирует Invoice по шаблону
        
        Returns:
            путь к сохраненному файлу
        """

        rows = []

        # Шапка документа
        seller_name_full = f"{metadata.seller_name} ({metadata.seller_name_en})" if metadata.seller_name_en else metadata.seller_name
        rows.append([seller_name_full])
        rows.append([metadata.seller_address])
        rows.append([metadata.seller_address_en if metadata.seller_address_en else ''])
        rows.append([f"ТЕЛ/TEL:  {self.preset['seller']['phone']}"])
        rows.append([''])  # Пустая строка
        rows.append([f"COMMERCIAL INVOICE / КОММЕРЧЕСКИЙ ИНВОЙС № {metadata.invoice_number} from/от {metadata.date}"])
        rows.append([''])  # Пустая строка
        rows.append([f"Buyer / Покупатель: {metadata.buyer_name}\n{metadata.buyer_address}"])
        rows.append([f"Contract / Контракт №{metadata.contract_number} from/от {metadata.contract_date}"])
        
        container_text = f"Terms of delivery / Условия поставки: {metadata.terms_of_delivery}"
        container_row = [container_text, '', '', '', '', '', '', '', '', f"Container No / Контейнер №", metadata.container_number]
        rows.append(container_row)
        
        rows.append([''])  # Пустая строка

        # Заголовок таблицы
        header = [
            "№",
            "Brand / Марка",
            "Code / Код ТНВЭД",
            "Factory code / Артикул",
            " Description / Описание",
            "Qty of pairs\nКол-во пар",
            "Net weight\nВес нетто, кг",
            "Gross weight\nВес брутто, кг",
            "Qty of places\nКол-во мест",
            "Price / Цена, cny",
            "Amount / \nСумма, cny"
        ]
        rows.append(header)

        # КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: БЕЗ ГРУППИРОВКИ - просто все строки подряд!
        total_qty = 0
        total_net = 0
        total_gross = 0
        total_boxes = 0
        total_amount = 0
        
        data_start_row = 13  # Строка, где начинаются данные (после заголовка)
        
        # Отслеживаем номер позиции и пары строк для объединения boxes
        item_number = 0
        prev_line = None

        for idx, line in enumerate(lines):
            # Определяем является ли строка "второй частью" артикула
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )
            
            # Увеличиваем номер только для первых строк
            if not is_continuation:
                item_number += 1
            
            # Формируем описание в зависимости от режима
            if self.mode == 'shipment':
                # Расширенное описание уже сформировано в ShipmentProcessor
                description = line.description
            else:
                # Старая логика для режима container
                description = f"{line.description}, материал верха: {line.material}, {line.insole_category}"

            rows.append([
                '' if is_continuation else item_number,  # Номер только для первой строки
                '' if is_continuation else line.brand,
                line.hs_code,
                '' if is_continuation else line.article,
                description,
                line.quantity,
                float(line.net_weight),
                float(line.gross_weight),
                '',  # Boxes - будет заполнено после объединения ячеек
                float(line.price),
                float(line.amount)
            ])

            total_qty += line.quantity
            total_net += line.net_weight
            total_gross += line.gross_weight
            # ИСПРАВЛЕНО: Правильный подсчет boxes
            # Для второй строки пары (is_continuation) НЕ добавляем boxes,
            # так как они уже были добавлены в первой строке
            if not is_continuation:
                total_boxes += line.boxes
            total_amount += line.amount
            
            prev_line = line

        data_end_row = len(rows)  # Последняя строка с данными

        # Итоговая строка
        rows.append([
            '',
            '',
            '',
            '',
            'Total / Итого:',
            total_qty,
            round(float(total_net), 3),
            round(float(total_gross), 3),
            total_boxes,
            '',
            f"¥{total_amount:,.2f}".replace(',', ' ')
        ])

        # Футер (только для Invoice - БЕЗ Terms of delivery и БЕЗ подписей!)
        rows.append([f"-Manufacturer / Производитель: {metadata.seller_name}"])
        rows.append([f"-Country of origin / Страна происхождения: {self.preset['delivery']['country_of_origin_en']} / {self.preset['delivery']['country_of_origin']}"])
        rows.append(["-Country of destination / Страна назначения: Russia / Россия"])
        rows.append(["-Product not for military use / Товар не для применения в военных целях"])
        rows.append(["-Terms of payment / Условия оплаты:"])
        rows.append([f"Payment of the cost of this transaction for the delivery of the goods specified above in the framework of the execution of Contract No. {metadata.contract_number} dated {metadata.contract_date} in the amount of ¥ {total_amount:,.2f} is payable no later than 120 days from the date of filing the Declaration for the goods in the country of Import./ Оплата стоимости данной сделки по поставке товара, указанного выше в рамках исполнения Контракта № {metadata.contract_number} от {metadata.contract_date} г. в размере ¥ {total_amount:,.2f} подлежит оплате не позднее 120 дней с даты подачи Декларации на товар в стране Импорта"])

        # Создаем DataFrame и сохраняем без форматирования
        df = pd.DataFrame(rows)
        # ВАЖНО: Используем контекстный менеджер чтобы файл корректно закрылся
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        
        # Объединяем ячейки boxes для пар строк
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Проходим по строкам данных и объединяем boxes для пар
        prev_line = None
        current_row = data_start_row
        
        for line in lines:
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )
            
            if is_continuation:
                # Объединяем ячейки boxes (колонка I) для предыдущей и текущей строки
                ws.merge_cells(f'I{current_row-1}:I{current_row}')
                # Записываем ИСХОДНОЕ значение boxes (сумма обеих строк)
                ws[f'I{current_row-1}'] = line.original_boxes  # ← ИСПРАВЛЕНО: используем original_boxes
            elif line.boxes > 0:
                # Для обычных строк просто записываем значение
                ws[f'I{current_row}'] = line.boxes
            
            current_row += 1
            prev_line = line
        
        wb.save(output_path)
        wb.close()
        
        # Применяем форматирование
        from ..formatters import InvoiceFormatter
        formatter = InvoiceFormatter(self.config)
        formatter.format(output_path, data_start_row, data_end_row)
        
        return output_path
