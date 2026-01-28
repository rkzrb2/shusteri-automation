"""
Генератор Packing List БЕЗ группировки
Выводит все строки из output_lines подряд
"""
import pandas as pd
from typing import List
from ..models import OutputLine, DocumentMetadata


class PackingListGenerator:
    """Генератор Упаковочного листа"""

    def __init__(self, config: dict, preset: dict, mode: str = 'container'):
        """
        Args:
            config: конфигурация
            preset: пресет клиента
            mode: режим работы ('container' или 'shipment')
        """
        self.config = config
        self.preset = preset
        self.mode = mode

    def generate(
            self,
            lines: List[OutputLine],
            metadata: DocumentMetadata,
            output_path: str
    ) -> str:
        """Генерирует Packing List"""

        rows = []

        # Шапка
        rows.append([f"{metadata.seller_name} ({self.preset['seller']['name_en']})"])
        rows.append([metadata.seller_address])
        rows.append([metadata.seller_address_en])
        rows.append([f"ТЕЛ/TEL:  {self.preset['seller']['phone']}"])
        rows.append([''])  # Пустая строка
        rows.append([f"Packing list / Упаковочный лист № {metadata.invoice_number} from/от {metadata.date}"])
        rows.append([''])  # Пустая строка
        rows.append([f"Buyer / Покупатель: {metadata.buyer_name}"])
        rows.append([metadata.buyer_address])
        rows.append([f"Contract / Контракт №{metadata.contract_number} from/от {metadata.contract_date}"])

        container_text = f"Terms of delivery / Условия поставки: {metadata.terms_of_delivery}"
        container_row = [container_text, '', '', '', '', '', '', '', '', f"Container No / Контейнер №", metadata.container_number]
        rows.append(container_row)

        rows.append([''])  # Пустая строка

        header = [
            "№",
            "Brand / Марка",
            "Code / Код ТНВЭД",
            "Factory code / Артикул",
            " Description / Описание",
            "Color / Цвет",
            "Quantity of pairs / Количество пар",
            "Net weight, kg / Вес нетто, кг",
            "Gross weight, kg / Вес брутто, кг",
            "Quantity of places / Количество мест",
            "Type of packaging / Вид упаковки"
        ]
        rows.append(header)

        # КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: БЕЗ ГРУППИРОВКИ - просто все строки подряд!
        total_qty = 0
        total_net = 0
        total_gross = 0
        total_boxes = 0
        
        data_start_row = 13  # Строка, где начинаются данные
        
        # Отслеживаем номер позиции и пары строк
        item_number = 0
        prev_line = None

        for idx, line in enumerate(lines):
            # Определяем является ли строка "второй частью" артикула
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )
            
            # Увеличиваем номер только для первых строк
            if not is_continuation:
                item_number += 1

            rows.append([
                '' if is_continuation else item_number,
                '' if is_continuation else line.brand,
                line.hs_code,
                '' if is_continuation else line.article,
                '' if is_continuation else line.description,
                '' if is_continuation else line.color,
                line.quantity,
                float(line.net_weight),
                float(line.gross_weight),
                '',  # Boxes - будет заполнено после объединения
                "cardboard box / \nкартонная коробка" if line.boxes > 0 or is_continuation else ''
            ])

            total_qty += line.quantity
            total_net += line.net_weight
            total_gross += line.gross_weight
            # ИСПРАВЛЕНО: Правильный подсчет boxes
            # Для второй строки пары (is_continuation) НЕ добавляем boxes,
            # так как они уже были добавлены в первой строке
            if not is_continuation:
                total_boxes += line.boxes
            
            prev_line = line

        data_end_row = len(rows)  # Последняя строка с данными

        # Итоги
        rows.append(['', '', '', '', '', 'Итого:', total_qty, round(float(total_net), 3), round(float(total_gross), 3), total_boxes, ''])
        rows.append([f"Total net weight, kgs / Общий вес нетто: {total_net:,.2f} кг"])
        rows.append([f"Total gross weight, kg / Общий вес брутто: {total_gross:,.2f} кг"])
        rows.append([f"Total quantity of pairs / Общее кол-во пар: {total_qty}"])
        rows.append([f"Total quantity of places / Общее кол-во мест: {total_boxes}"])

        # Создаем DataFrame и сохраняем
        df = pd.DataFrame(rows)
        # ВАЖНО: Используем контекстный менеджер чтобы файл корректно закрылся
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        
        # Объединяем ячейки boxes для пар строк
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Объединяем ячейки boxes для пар строк
        self._merge_boxes_container(ws, lines, data_start_row)
        
        wb.save(output_path)
        wb.close()
        
        # Применяем форматирование
        from ..formatters import PackingListFormatter
        formatter = PackingListFormatter(self.config)
        formatter.format(output_path, data_start_row, data_end_row)

        return output_path

    def _merge_boxes_container(self, ws, lines: List[OutputLine], data_start_row: int):
        """
        Объединяет ячейки boxes для пар строк с одним артикулом (≤24см и >24см)
        """
        prev_line = None
        current_row = data_start_row

        for line in lines:
            is_continuation = (
                prev_line is not None and
                line.article == prev_line.article and
                "более 24см" in line.insole_category and
                line.boxes == 0 and
                "до 24см" in prev_line.insole_category
            )

            if is_continuation:
                # Объединяем ячейки boxes (колонка J) для предыдущей и текущей строки
                ws.merge_cells(f'J{current_row-1}:J{current_row}')
                # Записываем ИСХОДНОЕ значение boxes (сумма обеих строк)
                ws[f'J{current_row-1}'] = line.original_boxes
            elif line.boxes > 0:
                # Для обычных строк просто записываем значение
                ws[f'J{current_row}'] = line.boxes

            current_row += 1
            prev_line = line
