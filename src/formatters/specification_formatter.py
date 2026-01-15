"""
Форматтер для Specification документа
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SpecificationFormatter:
    """Применяет форматирование к Specification"""
    
    def __init__(self, config: dict):
        self.config = config
    
    def format(self, file_path: str, data_start_row: int, data_end_row: int):
        """
        Применяет форматирование к файлу
        
        Args:
            file_path: путь к файлу Excel
            data_start_row: номер строки, где начинаются данные таблицы
            data_end_row: номер последней строки с данными
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Настройка ширины колонок
        self._set_column_widths(ws)
        
        # Форматирование шапки документа
        self._format_header(ws)
        
        # Форматирование таблицы
        self._format_table(ws, data_start_row, data_end_row)
        
        # Форматирование футера
        self._format_footer(ws, data_end_row)
        
        # Сохранение
        wb.save(file_path)
        wb.close()  # ВАЖНО: Закрываем файл чтобы Windows мог его удалить
    
    def _set_column_widths(self, ws):
        """Устанавливает ширину колонок"""
        widths = {
            'A': 3.33,   # №
            'B': 13.33,  # Brand
            'C': 12.0,   # Code
            'D': 18.5,   # Factory code
            'E': 21.0,   # Description
            'F': 14.0,   # Color
            'G': 22.0,   # Top material
            'H': 14.33,  # Lining material
            'I': 15.5,   # Outsole material
            'J': 12.16,  # Heel height
            'K': 15.33,  # Insole length
            'L': 12.16,  # Quantity
            'M': 15.0,   # Net weight
            'N': 18.83,  # Gross weight
            'O': 13.16,  # Quantity of places
            'P': 11.0,   # Price
            'Q': 15.0,   # Amount
            'R': 25.0    # KIZ codes
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _format_header(self, ws):
        """Форматирует шапку документа"""
        # Строки 1-3: Спецификация к Контракту (объединенные)
        for row in [1, 2, 3]:
            ws.merge_cells(f'A{row}:R{row}')
            cell = ws[f'A{row}']
            cell.font = Font(name='Arial', size=14, bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Строка 4: Container
        ws.merge_cells('A4:R4')
        cell = ws['A4']
        cell.font = Font(name='Arial', size=12)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Строка 5: пустая
        
        # Строка 6: пустая
        
        # Строка 7: пустая
        
        # Строка 8: Specification № ... from ...
        ws.merge_cells('A8:R8')
        cell = ws['A8']
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Строка 9: пустая
    
    def _format_table(self, ws, data_start_row, data_end_row):
        """Форматирует таблицу с данными"""
        # Заголовок таблицы (строка 10)
        header_row = 10
        ws.row_dimensions[header_row].height = 60
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков (18 колонок A-R)
        for col in range(1, 19):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Форматирование строк данных
        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, 19):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = Font(name='Arial', size=9)
                
                # Выравнивание для разных колонок
                if col == 1:  # №
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in [2, 3, 4]:  # Brand, Code, Article
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif col in [5, 6, 7, 8, 9, 10, 11]:  # Текстовые поля
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif col == 18:  # KIZ codes
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    def _format_footer(self, ws, data_end_row):
        """Форматирует футер документа"""
        total_row = data_end_row + 1
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Строка Total
        for col in range(1, 19):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Информационные строки после таблицы
        info_start = total_row + 1
        
        for row_offset in range(0, 9):  # 9 строк информации (включая подписи)
            row = info_start + row_offset
            ws.merge_cells(f'A{row}:R{row}')
            cell = ws[f'A{row}']
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Последние 2 строки для подписей (Buyer/Seller)
        signature_row = info_start + 9
        ws.merge_cells(f'A{signature_row}:I{signature_row}')
        ws.merge_cells(f'J{signature_row}:R{signature_row}')
        
        cell = ws[f'A{signature_row}']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        cell = ws[f'J{signature_row}']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Строка с названиями компаний
        company_row = signature_row + 1
        ws.merge_cells(f'A{company_row}:I{company_row}')
        ws.merge_cells(f'J{company_row}:R{company_row}')
        
        cell = ws[f'A{company_row}']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        cell = ws[f'J{company_row}']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
