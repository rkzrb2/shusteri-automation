"""
Форматтер для Invoice документа
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class InvoiceFormatter:
    """Применяет форматирование к Invoice"""
    
    def __init__(self, config: dict):
        self.config = config
    
    def format(self, file_path: str, data_start_row: int, data_end_row: int):
        """
        Применяет форматирование к файлу
        
        Args:
            file_path: путь к файлу Excel
            data_start_row: номер строки, где начинаются данные таблицы (после заголовка)
            data_end_row: номер последней строки с данными (перед Total)
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Настройка ширины колонок
        self._set_column_widths(ws)
        
        # Форматирование шапки документа
        self._format_header(ws)
        
        # Форматирование таблицы
        self._format_table(ws, data_start_row, data_end_row)
        
        # Форматирование футера
        self._format_footer(ws, data_end_row)
        
        # Сохранение
        wb.save(file_path)
        wb.close()  # ВАЖНО: Закрываем файл чтобы Windows мог его удалить
    
    def _set_column_widths(self, ws):
        """Устанавливает ширину колонок"""
        widths = {
            'A': 4.66,
            'B': 11.0,
            'C': 11.83,
            'D': 14.66,
            'E': 91.66,
            'F': 11.66,
            'G': 12.33,
            'H': 13.66,
            'I': 13.33,
            'J': 9.66,
            'K': 14.83
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _format_header(self, ws):
        """Форматирует шапку документа"""
        # Строка 1: Название продавца
        ws.merge_cells('A1:K1')
        cell = ws['A1']
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        ws.row_dimensions[1].height = 22.5
        
        # Строка 2: Адрес (строка 1)
        ws.merge_cells('A2:K2')
        cell = ws['A2']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 15.6
        
        # Строка 3: Адрес (строка 2)
        ws.merge_cells('A3:K3')
        cell = ws['A3']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 15.6
        
        # Строка 4: Телефон
        ws.merge_cells('A4:K4')
        cell = ws['A4']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 14.25
        
        # Строка 5: пустая
        ws.row_dimensions[5].height = 6.0
        
        # Строка 6: Заголовок документа COMMERCIAL INVOICE
        ws.merge_cells('A6:K6')
        cell = ws['A6']
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        ws.row_dimensions[6].height = 21.0
        
        # Строка 7: пустая
        ws.row_dimensions[7].height = 12.75
        
        # Строка 8: Buyer
        ws.merge_cells('A8:K8')
        cell = ws['A8']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[8].height = 49.5
        
        # Строка 9: Contract
        ws.merge_cells('A9:K9')
        cell = ws['A9']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[9].height = 14.25
        
        # Строка 10: Terms of delivery + Container
        ws.merge_cells('A10:I10')
        cell = ws['A10']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[10].height = 14.25
        
        # Строка 11: пустая
        ws.row_dimensions[11].height = 6.0
    
    def _format_table(self, ws, data_start_row, data_end_row):
        """Форматирует таблицу с данными"""
        # Заголовок таблицы (строка 12)
        header_row = 12
        ws.row_dimensions[header_row].height = 33.75
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for col in range(1, 12):  # A-K
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Форматирование строк данных
        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # Выравнивание для разных колонок
                if col == 1:  # №
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Arial', size=10)
                elif col in [2, 3, 4]:  # Brand, Code, Article
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(name='Arial', size=10)
                elif col == 5:  # Description
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(name='Arial', size=10)
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.font = Font(name='Arial', size=10)
    
    def _format_footer(self, ws, data_end_row):
        """Форматирует футер документа"""
        total_row = data_end_row + 1
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Строка Total
        for col in range(1, 12):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10, bold=True)
            
            if col == 5:  # "Total / Итого:"
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in [6, 7, 8, 9]:  # Числовые поля
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 11:  # Amount
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Информационные строки после таблицы
        info_start = total_row + 1
        
        for row_offset in range(0, 7):  # 7 строк информации
            row = info_start + row_offset
            ws.merge_cells(f'A{row}:K{row}')
            cell = ws[f'A{row}']
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
