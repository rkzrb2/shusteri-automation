"""
Форматтер для Packing List документа
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


class PackingListFormatter:
    """Применяет форматирование к Packing List"""
    
    def __init__(self, config: dict):
        self.config = config
    
    def format(self, file_path: str, data_start_row: int, data_end_row: int):
        """
        Применяет форматирование к файлу
        
        Args:
            file_path: путь к файлу Excel
            data_start_row: номер строки, где начинаются данные таблицы
            data_end_row: номер последней строки с данными
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Настройка ширины колонок
        self._set_column_widths(ws)
        
        # Форматирование шапки документа
        self._format_header(ws)
        
        # Форматирование таблицы
        self._format_table(ws, data_start_row, data_end_row)
        
        # Форматирование футера
        self._format_footer(ws, data_end_row)
        
        # Сохранение
        wb.save(file_path)
        wb.close()  # ВАЖНО: Закрываем файл чтобы Windows мог его удалить
    
    def _set_column_widths(self, ws):
        """Устанавливает ширину колонок"""
        widths = {
            'A': 3.33,   # №
            'B': 13.5,   # Brand
            'C': 13.33,  # Code
            'D': 16.0,   # Factory code
            'E': 27.0,   # Description
            'F': 21.0,   # Color
            'G': 17.16,  # Quantity
            'H': 17.5,   # Net weight
            'I': 19.66,  # Gross weight
            'J': 18.66,  # Quantity of places
            'K': 21.16   # Type of packaging
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _format_header(self, ws):
        """Форматирует шапку документа"""
        # Строка 1: Название продавца
        ws.merge_cells('A1:K1')
        cell = ws['A1']
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        ws.row_dimensions[1].height = 22.5
        
        # Строка 2: Адрес (строка 1)
        ws.merge_cells('A2:K2')
        cell = ws['A2']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 15.6
        
        # Строка 3: Адрес (строка 2)
        ws.merge_cells('A3:K3')
        cell = ws['A3']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 15.6
        
        # Строка 4: Телефон
        ws.merge_cells('A4:K4')
        cell = ws['A4']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 14.25
        
        # Строка 5: пустая
        ws.row_dimensions[5].height = 6.0
        
        # Строка 6: Заголовок документа Packing list
        ws.merge_cells('A6:K6')
        cell = ws['A6']
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        ws.row_dimensions[6].height = 21.0
        
        # Строка 7: пустая
        ws.row_dimensions[7].height = 12.75
        
        # Строка 8: Buyer
        ws.merge_cells('A8:K8')
        cell = ws['A8']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[8].height = 14.25
        
        # Строка 9: Адрес покупателя
        ws.merge_cells('A9:K9')
        cell = ws['A9']
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[9].height = 49.5
        
        # Строка 10: Contract
        ws.merge_cells('A10:K10')
        cell = ws['A10']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[10].height = 14.25
        
        # Строка 11: Terms of delivery + Container
        ws.merge_cells('A11:I11')
        cell = ws['A11']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[11].height = 14.25
        
        # Строка 12: пустая
        ws.row_dimensions[12].height = 6.0
    
    def _format_table(self, ws, data_start_row, data_end_row):
        """Форматирует таблицу с данными"""
        # Заголовок таблицы (строка 13)
        header_row = 13
        ws.row_dimensions[header_row].height = 50
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for col in range(1, 12):  # A-K
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Форматирование строк данных
        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = Font(name='Arial', size=10)
                
                # Выравнивание для разных колонок
                if col == 1:  # №
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in [2, 3, 4]:  # Brand, Code, Article
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif col in [5, 6]:  # Description, Color
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif col == 11:  # Type of packaging
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    def _format_footer(self, ws, data_end_row):
        """Форматирует футер документа"""
        total_row = data_end_row + 1
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Строка Total (Итого:)
        for col in range(1, 12):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10, bold=True)
            
            if col == 6:  # "Итого:"
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in [7, 8, 9, 10]:  # Числовые поля
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Информационные строки после таблицы (4 строки)
        info_start = total_row + 1
        
        for row_offset in range(0, 4):
            row = info_start + row_offset
            ws.merge_cells(f'A{row}:K{row}')
            cell = ws[f'A{row}']
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
