"""
Инъектор КИЗ кодов в уже сгенерированный файл Спецификации.

Позволяет добавить коды маркировки (Честный знак) в колонку КИЗ
без полной перегенерации документов.
"""
import logging
from pathlib import Path
from openpyxl import load_workbook

from .km_loader import KMLoader

logger = logging.getLogger(__name__)

# Колонки в листе Спецификации (1-based)
_COL_ITEM_NUM = 1   # A: №
_COL_ARTICLE  = 4   # D: Factory code / Артикул
_COL_INSOLE   = 11  # K: Insole length / Длина стельки
_COL_KIZ      = 18  # R: КИЗ

_SPEC_SHEET_NAMES = ("Specification", "Спецификация", "specification")


class KIZInjector:
    """Добавляет КИЗ коды в колонку R листа Спецификации существующего файла."""

    def __init__(self, km_loader: KMLoader):
        self.km_loader = km_loader

    def inject(self, target_file: Path, output_file: Path = None) -> dict:
        """
        Записывает КИЗ коды в колонку КИЗ (R) листа Спецификации.

        Форматирование ячеек сохраняется — только значение КИЗ обновляется.

        Args:
            target_file: путь к существующему .xlsx файлу (combined 3-sheet или отдельная Spec)
            output_file: путь к файлу-результату; если None — перезаписывает target_file

        Returns:
            {'rows_updated': int, 'codes_added': int, 'rows_skipped': int}
        """
        if output_file is None:
            output_file = target_file

        wb = load_workbook(str(target_file))
        ws = self._find_spec_sheet(wb, target_file)

        rows_updated = 0
        codes_added = 0
        rows_skipped = 0

        for row in ws.iter_rows():
            # Строка данных — только если в колонке A стоит целое число (номер позиции)
            cell_a = row[_COL_ITEM_NUM - 1]
            if not isinstance(cell_a.value, (int, float)) or cell_a.value != int(cell_a.value):
                continue

            article_cell = row[_COL_ARTICLE - 1]
            insole_cell  = row[_COL_INSOLE - 1]

            article    = str(article_cell.value).strip() if article_cell.value else ""
            insole_cat = str(insole_cell.value).strip()  if insole_cell.value  else ""

            if not article:
                rows_skipped += 1
                continue

            km_codes = self.km_loader.get_km_codes_for_category(article, insole_cat)

            kiz_cell = row[_COL_KIZ - 1]
            if km_codes:
                kiz_cell.value = "\n".join(km_codes)
                rows_updated += 1
                codes_added  += len(km_codes)
                logger.info(f"  {article} ({insole_cat}): {len(km_codes)} кодов")
            else:
                logger.warning(f"  Нет КМ кодов для {article} ({insole_cat})")
                rows_skipped += 1

        wb.save(str(output_file))
        wb.close()

        return {
            "rows_updated": rows_updated,
            "codes_added":  codes_added,
            "rows_skipped": rows_skipped,
        }

    # ------------------------------------------------------------------
    def _find_spec_sheet(self, wb, target_file: Path):
        """Находит лист Спецификации в workbook."""
        # Точное совпадение по известным именам
        for name in _SPEC_SHEET_NAMES:
            if name in wb.sheetnames:
                logger.info(f"Найден лист Спецификации: '{name}'")
                return wb[name]

        # Нечёткий поиск: имя содержит 'spec' или 'спец'
        for sheet_name in wb.sheetnames:
            if "spec" in sheet_name.lower() or "спец" in sheet_name.lower():
                logger.info(f"Найден лист Спецификации (нечёткий поиск): '{sheet_name}'")
                return wb[sheet_name]

        # Единственный лист — используем его
        if len(wb.sheetnames) == 1:
            logger.info(f"Один лист в файле, используем: '{wb.sheetnames[0]}'")
            return wb.active

        raise ValueError(
            f"Не найден лист Спецификации в '{target_file.name}'.\n"
            f"Доступные листы: {', '.join(wb.sheetnames)}"
        )
