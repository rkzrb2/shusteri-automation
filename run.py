#!/usr/bin/env python3
"""
Автоматизация документооборота для Шустери
Генерация Invoice, Specification и Packing List с форматированием

Интерактивный режим работы
"""

import yaml
from pathlib import Path
from datetime import datetime
import logging
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt, Confirm

from src.parser import InputFileParser
from src.processor import DataProcessor
from src.shipment_parser import ShipmentParser
from src.shipment_processor import ShipmentProcessor
from src.generators.invoice import InvoiceGenerator
from src.generators.specification import SpecificationGenerator
from src.generators.packing_list import PackingListGenerator
from src.models import DocumentMetadata
from src.km_loader import KMLoader

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('automation.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

console = Console()


class ShusteriAutomation:
    """Главный класс автоматизации"""

    def __init__(self, config_path: str = "config.yaml"):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
        
        self.preset = None  # Текущий выбранный пресет
    
    def get_available_presets(self):
        """Получает список доступных пресетов"""
        presets_dir = Path("presets")
        if not presets_dir.exists():
            console.print("[bold red]❌ Папка presets/ не найдена![/bold red]")
            console.print("[yellow]Создайте папку presets/ и положите туда файлы пресетов[/yellow]")
            return []
        
        # Ищем YAML файлы
        preset_files = list(presets_dir.glob("*.yaml")) + list(presets_dir.glob("*.yml"))
        
        if not preset_files:
            console.print("[bold red]❌ Нет пресетов в папке presets/![/bold red]")
            return []
        
        # Загружаем информацию о пресетах
        presets = []
        for preset_file in sorted(preset_files):
            try:
                with open(preset_file, 'r', encoding='utf-8') as f:
                    preset_data = yaml.safe_load(f)
                    presets.append({
                        'file': preset_file,
                        'name': preset_data.get('preset_name', preset_file.stem),
                        'description': preset_data.get('description', ''),
                        'data': preset_data
                    })
            except Exception as e:
                logger.warning(f"Не удалось загрузить пресет {preset_file}: {e}")
        
        return presets
    
    def select_preset(self):
        """Интерактивный выбор пресета"""
        presets = self.get_available_presets()
        
        if not presets:
            return None
        
        # Показываем список пресетов
        console.print("\n[bold cyan]👥 Доступные пресеты клиентов:[/bold cyan]\n")
        
        table = Table(show_header=True, header_style="bold magenta")
        table.add_column("№", style="cyan", width=4)
        table.add_column("Клиент", style="green")
        table.add_column("Описание", style="yellow")
        
        for idx, preset in enumerate(presets, 1):
            table.add_row(
                str(idx),
                preset['name'],
                preset['description']
            )
        
        console.print(table)
        console.print()
        
        # Запрашиваем выбор
        if len(presets) == 1:
            console.print(f"[green]✓ Автоматически выбран пресет: {presets[0]['name']}[/green]\n")
            return presets[0]['data']
        else:
            choice = Prompt.ask(
                "Выберите пресет (введите номер)",
                choices=[str(i) for i in range(1, len(presets) + 1)],
                default="1"
            )
            selected_preset = presets[int(choice) - 1]
            console.print(f"[green]✓ Выбран пресет: {selected_preset['name']}[/green]\n")
            return selected_preset['data']

    def get_input_files(self):
        """Получает список файлов в папке input/"""
        input_dir = Path("input")
        if not input_dir.exists():
            console.print("[bold red]❌ Папка input/ не найдена![/bold red]")
            console.print("[yellow]Создайте папку input/ и положите туда Excel файлы[/yellow]")
            return []
        
        # Ищем Excel файлы
        excel_files = list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls"))
        excel_files = [f for f in excel_files if not f.name.startswith('~')]  # Исключаем временные файлы
        
        return sorted(excel_files)

    def select_input_file(self):
        """Интерактивный выбор входного файла"""
        files = self.get_input_files()
        
        if not files:
            console.print("[bold red]❌ Нет Excel файлов в папке input/![/bold red]")
            console.print("[yellow]Положите Excel файл в папку input/ и запустите программу снова[/yellow]")
            return None
        
        # Показываем список файлов
        console.print("\n[bold cyan]📁 Доступные файлы:[/bold cyan]\n")
        
        table = Table(show_header=True, header_style="bold magenta")
        table.add_column("№", style="cyan", width=4)
        table.add_column("Имя файла", style="green")
        table.add_column("Размер", style="yellow", justify="right")
        table.add_column("Дата изменения", style="blue")
        
        for idx, file in enumerate(files, 1):
            size_kb = file.stat().st_size / 1024
            mtime = datetime.fromtimestamp(file.stat().st_mtime).strftime('%d.%m.%Y %H:%M')
            table.add_row(str(idx), file.name, f"{size_kb:.1f} KB", mtime)
        
        console.print(table)
        console.print()
        
        # Запрашиваем выбор
        if len(files) == 1:
            # Если файл один - используем его автоматически
            console.print(f"[green]✓ Автоматически выбран файл: {files[0].name}[/green]\n")
            return files[0]
        else:
            # Если файлов несколько - даем выбрать
            choice = Prompt.ask(
                "Выберите файл (введите номер)",
                choices=[str(i) for i in range(1, len(files) + 1)],
                default="1"
            )
            selected_file = files[int(choice) - 1]
            console.print(f"[green]✓ Выбран файл: {selected_file.name}[/green]\n")
            return selected_file

    def get_km_files(self):
        """Получает список файлов в папке выгрузка честный знак/"""
        km_dir = Path("выгрузка честный знак")
        if not km_dir.exists():
            return []

        # Ищем Excel файлы
        excel_files = list(km_dir.glob("*.xlsx")) + list(km_dir.glob("*.xls"))
        excel_files = [f for f in excel_files if not f.name.startswith('~')]  # Исключаем временные файлы

        return sorted(excel_files)

    def select_km_file(self):
        """Интерактивный выбор файла маркировки (КМ)"""
        # Спрашиваем, нужен ли файл КМ
        use_km = Confirm.ask(
            "\n🏷️  Использовать файл маркировки (Честный знак)?",
            default=False
        )

        if not use_km:
            return None

        files = self.get_km_files()

        if not files:
            console.print("[yellow]⚠️  Нет файлов в папке 'выгрузка честный знак/'[/yellow]")
            console.print("[yellow]   Продолжаем без маркировки[/yellow]\n")
            return None

        # Показываем список файлов
        console.print("\n[bold cyan]📁 Доступные файлы маркировки:[/bold cyan]\n")

        table = Table(show_header=True, header_style="bold magenta")
        table.add_column("№", style="cyan", width=4)
        table.add_column("Имя файла", style="green")
        table.add_column("Размер", style="yellow", justify="right")
        table.add_column("Дата изменения", style="blue")

        for idx, file in enumerate(files, 1):
            size_kb = file.stat().st_size / 1024
            mtime = datetime.fromtimestamp(file.stat().st_mtime).strftime('%d.%m.%Y %H:%M')
            table.add_row(str(idx), file.name, f"{size_kb:.1f} KB", mtime)

        console.print(table)
        console.print()

        # Запрашиваем выбор
        if len(files) == 1:
            console.print(f"[green]✓ Автоматически выбран файл: {files[0].name}[/green]\n")
            return files[0]
        else:
            choice = Prompt.ask(
                "Выберите файл (введите номер)",
                choices=[str(i) for i in range(1, len(files) + 1)],
                default="1"
            )
            selected_file = files[int(choice) - 1]
            console.print(f"[green]✓ Выбран файл: {selected_file.name}[/green]\n")
            return selected_file

    def enrich_with_km_codes(self, output_lines, km_loader: KMLoader):
        """
        Обогащает output_lines кодами маркировки из справочника КМ.

        Новая логика:
        - Использует точное количество кодов для каждого размера (по qty_by_size)
        - Отслеживает использованные коды, чтобы для артикулов с PRG и без PRG
          выдавались РАЗНЫЕ коды (последовательно из справочника)

        Args:
            output_lines: список OutputLine
            km_loader: загруженный справочник КМ
        """
        enriched_count = 0
        total_km_codes = 0

        for line in output_lines:
            # Проверяем, есть ли информация о размерах
            if not line.qty_by_size:
                logger.warning(f"Нет информации о размерах для артикула {line.article}, пропускаем")
                continue

            # Получаем ТОЧНОЕ количество КМ кодов для конкретных размеров
            km_codes = km_loader.get_km_codes_exact(line.article, line.qty_by_size)

            if km_codes:
                line.kiz_codes = km_codes
                enriched_count += 1
                total_km_codes += len(km_codes)

        logger.info(f"Обогащено {enriched_count} строк, всего {total_km_codes} КМ кодов")
        return enriched_count, total_km_codes

    def get_invoice_number(self):
        """Запрос номера инвойса"""
        # Генерируем номер по умолчанию на основе текущей даты
        default_number = datetime.now().strftime('%Y%m%d01')
        
        invoice_number = Prompt.ask(
            "📋 Введите номер инвойса",
            default=default_number
        )
        
        return invoice_number

    def get_container_number(self):
        """Запрос номера контейнера (опционально)"""
        has_container = Confirm.ask(
            "📦 Есть номер контейнера?",
            default=False
        )
        
        if has_container:
            container = Prompt.ask("Введите номер контейнера")
            return container
        else:
            return ""
    
    def select_processing_mode(self):
        """Выбор режима обработки"""
        console.print("\n[bold cyan]⚙️  Выберите тип обработки:[/bold cyan]\n")
        console.print("  [yellow]1[/yellow] - Загрузка контейнера (стандартный формат)")
        console.print("      • 1 товар = 1 строка со всеми размерами (35-42)")
        console.print("      • Единица измерения: пары")
        console.print()
        console.print("  [yellow]2[/yellow] - Отправка грузов (новый формат - полупары)")
        console.print("      • 1 товар 1 размера = 1 строка")
        console.print("      • Единица измерения: полупары (левый/правый)")
        console.print("      • Расширенное описание товара")

        choice = Prompt.ask(
            "\nВведите номер",
            choices=["1", "2"],
            default="1"
        )

        mode = 'container' if choice == '1' else 'shipment'
        mode_name = "Загрузка контейнера" if mode == 'container' else "Отправка грузов (полупары)"
        console.print(f"[green]✓ Выбран режим: {mode_name}[/green]\n")

        return mode

    def get_output_format(self):
        """Запрос формата вывода"""
        console.print("\n[cyan]📄 Выберите формат вывода:[/cyan]")
        console.print("  [yellow]1[/yellow] - Три отдельных файла (Invoice.xlsx, Specification.xlsx, PackingList.xlsx)")
        console.print("  [yellow]2[/yellow] - Один файл с тремя листами (All_Documents.xlsx)")

        choice = Prompt.ask(
            "\nВыберите вариант",
            choices=["1", "2"],
            default="1"
        )

        return choice
    
    def generate_combined_file(
            self,
            output_lines,
            metadata,
            output_path: Path,
            base_name: str,
            mode: str = 'container'
    ):
        """Генерирует один файл с тремя листами"""
        from openpyxl import Workbook, load_workbook
        import tempfile
        import os

        combined_file = output_path / f"{base_name}_All_Documents.xlsx"

        # Создаем временные файлы для каждого документа
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            temp_invoice = os.path.join(tmpdir, "invoice.xlsx")
            temp_spec = os.path.join(tmpdir, "spec.xlsx")
            temp_packing = os.path.join(tmpdir, "packing.xlsx")

            # Генерируем каждый документ во временные файлы
            invoice_gen = InvoiceGenerator(self.config, self.preset, mode=mode)
            invoice_gen.generate(output_lines, metadata, temp_invoice)

            spec_gen = SpecificationGenerator(self.config, self.preset, mode=mode)
            spec_gen.generate(output_lines, metadata, temp_spec)

            packing_gen = PackingListGenerator(self.config, self.preset, mode=mode)
            packing_gen.generate(output_lines, metadata, temp_packing)
            
            # Создаем новый workbook для объединения
            wb_combined = Workbook()
            wb_combined.remove(wb_combined.active)  # Удаляем дефолтный лист
            
            # Копируем лист Invoice
            wb_invoice = load_workbook(temp_invoice)
            ws_invoice = wb_invoice.active
            ws_new_invoice = wb_combined.create_sheet("Invoice")
            self._copy_sheet(ws_invoice, ws_new_invoice)
            wb_invoice.close()
            
            # Копируем лист Specification
            wb_spec = load_workbook(temp_spec)
            ws_spec = wb_spec.active
            ws_new_spec = wb_combined.create_sheet("Specification")
            self._copy_sheet(ws_spec, ws_new_spec)
            wb_spec.close()
            
            # Копируем лист Packing List
            wb_packing = load_workbook(temp_packing)
            ws_packing = wb_packing.active
            ws_new_packing = wb_combined.create_sheet("Packing List")
            self._copy_sheet(ws_packing, ws_new_packing)
            wb_packing.close()
            
            # Сохраняем объединенный файл
            wb_combined.save(str(combined_file))
            wb_combined.close()
        
        return combined_file
    
    def _copy_sheet(self, source_ws, target_ws):
        """Копирует содержимое и форматирование из одного листа в другой"""
        from copy import copy
        
        # Копируем размеры колонок
        for column in source_ws.column_dimensions:
            target_ws.column_dimensions[column].width = source_ws.column_dimensions[column].width
        
        # Копируем размеры строк
        for row in source_ws.row_dimensions:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
        
        # Копируем ячейки
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws[cell.coordinate]
                
                # Копируем значение
                target_cell.value = cell.value
                
                # Копируем форматирование
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = copy(cell.number_format)
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)
        
        # Копируем объединенные ячейки
        for merged_cell_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_cell_range))

    def process(
            self,
            input_file: Path,
            invoice_number: str,
            mode: str = 'container',
            container_number: str = "",
            output_format: str = "1",
            output_dir: str = "output",
            km_file: Path = None
    ):
        """Главный метод обработки"""

        console.print(f"\n[bold blue]🚀 Начало обработки Invoice #{invoice_number}[/bold blue]\n")

        try:
            # 1. Парсинг входного файла в зависимости от режима
            console.print("[cyan]📥 Чтение входного файла...[/cyan]")

            if mode == 'container':
                # Старый формат: загрузка контейнера
                parser = InputFileParser(self.config)
                products = parser.parse(str(input_file))

                if not products:
                    console.print("[bold red]❌ Не найдено ни одной позиции в файле[/bold red]")
                    return

                console.print(f"[green]✓ Загружено {len(products)} позиций[/green]\n")

                # Обработка данных
                console.print("[cyan]⚙️  Обработка данных...[/cyan]")
                processor = DataProcessor(self.config)
                output_lines = processor.process(products)

            else:
                # Новый формат: отправка грузов (полупары)
                parser = ShipmentParser(str(input_file))
                shipment_lines = parser.parse()

                if not shipment_lines:
                    console.print("[bold red]❌ Не найдено ни одной позиции в файле[/bold red]")
                    return

                console.print(f"[green]✓ Загружено {len(shipment_lines)} строк (полупар)[/green]\n")

                # Обработка данных
                console.print("[cyan]⚙️  Обработка данных...[/cyan]")
                processor = ShipmentProcessor(self.config)
                output_lines = processor.process(shipment_lines)

            console.print(f"[green]✓ Создано {len(output_lines)} строк для документов[/green]\n")

            # 2.5. Обогащение данных кодами маркировки (КМ)
            if km_file:
                console.print("[cyan]🏷️  Загрузка кодов маркировки...[/cyan]")
                try:
                    km_loader = KMLoader(str(km_file))
                    enriched_count, total_km = self.enrich_with_km_codes(output_lines, km_loader)
                    console.print(f"[green]✓ Добавлено {total_km} КМ кодов в {enriched_count} строк[/green]\n")
                except Exception as e:
                    logger.error(f"Ошибка загрузки КМ: {e}")
                    console.print(f"[yellow]⚠️  Ошибка загрузки КМ: {e}[/yellow]")
                    console.print("[yellow]   Продолжаем без маркировки[/yellow]\n")

            # 3. Подготовка метаданных
            date = datetime.now().strftime('%d.%m.%Y')
            metadata = DocumentMetadata(
                invoice_number=invoice_number,
                date=date,
                container_number=container_number,
                seller_name=self.preset['seller']['name'],
                seller_name_en=self.preset['seller'].get('name_en', ''),
                seller_address=self.preset['seller']['address'],
                seller_address_en=self.preset['seller'].get('address_en', ''),
                buyer_name=self.preset['buyer']['name'],
                buyer_address=self.preset['buyer']['address'],
                buyer_address_en=self.preset['buyer'].get('address_en', ''),
                contract_number=self.preset['contract']['number'],
                contract_date=self.preset['contract']['date'],
                terms_of_delivery=self.preset['delivery']['terms'],
                currency=self.preset['delivery']['currency']
            )

            # 4. Создание директории для выходных файлов
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)

            # Имена файлов
            base_name = f"Shusteri_{invoice_number}_{datetime.now().strftime('%Y%m%d')}"

            # 5. Генерация документов с форматированием
            console.print("[cyan]📄 Генерация документов с форматированием...[/cyan]")

            if output_format == "1":
                # Три отдельных файла
                invoice_file = output_path / f"{base_name}_Invoice.xlsx"
                spec_file = output_path / f"{base_name}_Specification.xlsx"
                packing_file = output_path / f"{base_name}_PackingList.xlsx"

                invoice_gen = InvoiceGenerator(self.config, self.preset, mode=mode)
                invoice_gen.generate(output_lines, metadata, str(invoice_file))

                spec_gen = SpecificationGenerator(self.config, self.preset, mode=mode)
                spec_gen.generate(output_lines, metadata, str(spec_file))

                packing_gen = PackingListGenerator(self.config, self.preset, mode=mode)
                packing_gen.generate(output_lines, metadata, str(packing_file))
                
                generated_files = [
                    ("Invoice", invoice_file.name),
                    ("Specification", spec_file.name),
                    ("Packing List", packing_file.name)
                ]
            else:
                # Один файл с тремя листами
                combined_file = self.generate_combined_file(
                    output_lines,
                    metadata,
                    output_path,
                    base_name,
                    mode
                )
                
                generated_files = [
                    ("All Documents (3 sheets)", combined_file.name)
                ]

            console.print("[green]✓ Документы сгенерированы с форматированием[/green]\n")

            # Вывод результатов
            console.print(f"\n[bold green]✅ Обработка завершена успешно![/bold green]\n")

            result_table = Table(title="Сгенерированные файлы")
            result_table.add_column("Документ", style="cyan")
            result_table.add_column("Файл", style="green")

            for doc_name, file_name in generated_files:
                result_table.add_row(doc_name, file_name)

            console.print(result_table)

            # Статистика
            total_qty = sum(line.quantity for line in output_lines)
            total_amount = sum(line.amount for line in output_lines)
            total_net = sum(line.net_weight for line in output_lines)
            total_gross = sum(line.gross_weight for line in output_lines)

            stats_table = Table(title="Статистика")
            stats_table.add_column("Параметр", style="cyan")
            stats_table.add_column("Значение", style="magenta")

            # Определяем количество позиций в зависимости от режима
            if mode == 'container':
                positions_count = len(products)
                unit_label = "Всего пар"
            else:
                positions_count = len(shipment_lines)
                unit_label = "Всего полупар"

            stats_table.add_row(unit_label, str(total_qty))
            stats_table.add_row("Сумма (CNY)", f"¥{total_amount:,.2f}")
            stats_table.add_row("Вес нетто (кг)", f"{total_net:,.3f}")
            stats_table.add_row("Вес брутто (кг)", f"{total_gross:,.3f}")
            stats_table.add_row("Позиций", str(positions_count))
            stats_table.add_row("Строк в документах", str(len(output_lines)))

            console.print(stats_table)

            console.print(f"\n[bold cyan]📁 Файлы сохранены в: {output_path.absolute()}[/bold cyan]\n")
            
            # Предложение создать еще один документ
            another = Confirm.ask("\n💡 Создать еще один документ?", default=False)
            return another

        except Exception as e:
            logger.error(f"Критическая ошибка: {e}", exc_info=True)
            console.print(f"\n[bold red]💥 Ошибка: {e}[/bold red]\n")
            raise


def main():
    """
    Интерактивный режим работы
    """
    console.print("\n[bold blue]═══════════════════════════════════════════════[/bold blue]")
    console.print("[bold blue]   Shusteri Automation - Генерация документов   [/bold blue]")
    console.print("[bold blue]═══════════════════════════════════════════════[/bold blue]\n")
    
    try:
        automation = ShusteriAutomation()
        
        # Выбор пресета клиента
        preset = automation.select_preset()
        if not preset:
            console.print("[bold red]Не удалось выбрать пресет. Завершение работы.[/bold red]")
            return
        
        automation.preset = preset
        
        # Цикл для создания нескольких документов
        while True:
            # 1. Выбор режима обработки
            mode = automation.select_processing_mode()

            # 2. Выбор входного файла
            input_file = automation.select_input_file()
            if not input_file:
                break

            # 3. Запрос номера инвойса
            invoice_number = automation.get_invoice_number()

            # 4. Запрос номера контейнера
            container_number = automation.get_container_number()

            # 5. Запрос формата вывода
            output_format = automation.get_output_format()

            # 6. Выбор файла маркировки (опционально)
            km_file = automation.select_km_file()

            # 7. Обработка
            create_another = automation.process(
                input_file,
                invoice_number,
                mode,
                container_number,
                output_format,
                km_file=km_file
            )
            
            # Если пользователь не хочет создавать еще документы - выходим
            if not create_another:
                break
        
        console.print("\n[bold green]👋 До свидания![/bold green]\n")

    except KeyboardInterrupt:
        console.print("\n\n[yellow]⚠️  Прервано пользователем[/yellow]\n")
    except Exception as e:
        console.print(f"\n[bold red]💥 Критическая ошибка: {e}[/bold red]\n")
        exit(1)


if __name__ == "__main__":
    main()
